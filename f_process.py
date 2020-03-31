#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Nov 29 15:47:02 2018

@author: ashkar
"""
#import built-in modules
import re
import pandas as pd
import os
import shutil
import time
import json
import xlrd
import openpyxl
import smtplib

from email.message import EmailMessage

#import manual modules
import config
import datatypes
import regex
import errors
import mail
import f_load

def isLeapYear(year):
    
    if ((year % 100 == 0 and year % 400 == 0) or (year % 100 != 0 and year % 4 == 0)):
        return True
    return False
    
def assignTypes(format_type, sub_type, header_type, index, formatType_element, subType_element, headerType_element):
    
    format_type[index] = formatType_element
    sub_type[index] = subType_element
    header_type[index] = headerType_element
    
#def isCellEmpty(cellVal):
#    if len(cellVal) == 0 or cellVal == 'nan' or cellVal == 'nat' or cellVal == 'na':
#        return True
#    return False

def isCellEmpty(cellVal):
    if len(cellVal) == 0 or cellVal == 'nan' or cellVal == 'nat':
        return True
    return False

def isEveryElementEmpty(lst):
    temp_bool = True
    
    for i in lst:
        temp = i.replace(" ", "")
        if not isCellEmpty(temp):
            temp_bool = False
            break
    
    return [temp_bool]

def isExistsDuplicate(lst, deletedHeaders_asListRev, first_data_index, user_correction_status):
    '''Returns True if duplicate headers exist'''
    if user_correction_status == config.USER_CORRECTION_STATUS['success']:
        lst = neglect_indices(lst, deletedHeaders_asListRev)
        
    temp_bool_dup = False
    temp_bool_empty = False
    distinct_list = []
    length_of_list = len(lst)
    
    for i in range(first_data_index, length_of_list):
        temp = lst[i].replace(" ", "")
        temp_lowercase = temp.lower()
        
        cell_empty = isCellEmpty(temp_lowercase)
        if cell_empty:
            temp_bool_empty = True
            break
        else:
            if (not cell_empty and temp_lowercase not in distinct_list):
                
                distinct_list.append(temp_lowercase)
            
            elif (not cell_empty and temp_lowercase in distinct_list):
                
                temp_bool_dup = True
                break    
        
        # if "\n" in lst[i]:
        #     lst[i].replace("\n", " ")
        # lst[i].strip(" ", "")

    return [temp_bool_dup, lst[first_data_index::], temp_bool_empty]

def generateLists(lst):
    header_type = ["not updated" for i in range(len(lst))]
    sub_type = ["sub not updated" for i in range(len(lst))]
    format_type = ["format not updated" for i in range(len(lst))]
    chop_zero = [0 for i in range(len(lst))]
    
    return [header_type, sub_type, format_type, chop_zero]

def isExistsSpecificCurrencySymbolInNumber(temp):
    tempBool = False
    
    for currency_symbol in datatypes.CURRENCY_SYMBOLS:
        if currency_symbol in temp:
            tempBool = True
            break
        
    return tempBool


######### NEW METHODS START ######################################

def cleanNumberCurrency(temp):

    list_unnecessary = datatypes.CURRENCY_SYMBOLS + ["rs.", ",", "(", ")"]

    for symbol in list_unnecessary:
        if symbol in temp:
            temp = temp.replace(symbol, "")

    return temp

def cleanNumberPercentage(temp):   
    
    list_unnecessary = ["%", "(", ")"]

    for symbol in list_unnecessary:
        if symbol in temp:
            temp = temp.replace(symbol, "")

    return temp

def cleanNumberNumber(temp):
    
    list_unnecessary = ['#div/0!', ",", "(", ")"]

    for symbol in list_unnecessary:
        if symbol in temp:
            temp = temp.replace(symbol, "")
        
    return temp

def changeDateSeparatorToSlash(tempDate):
    
    if not re.search(r'(\d+ \d+:)', tempDate):#execute if subtype is not datetime
        
        list_date_separator = ["-", " "]
        
        for date_separator in list_date_separator:
            if date_separator in tempDate:
                tempDate = tempDate.replace(date_separator, "/")
                
    else:#execute if subtype is datetime
        if re.search(r'(\d+:)', tempDate.split(" ")[1]):#if " " is not the date separator
            if "-" in tempDate:
                tempDate = tempDate.replace("-", "/")
                
        else:
            tempDate_splitted = tempDate.split(" ")
            if tempDate.count(" ") == 3:#if " " is the date separator
                tempDate = tempDate_splitted[0] + "/" + tempDate_splitted[1] + "/" + tempDate_splitted[2] + " " + tempDate_splitted[3]
                
            elif tempDate.count(" ") == 2:#if any one of the two date separator is " "
                tempDate = tempDate_splitted[0] + "/" + tempDate_splitted[1] + " " + tempDate_splitted[2]
                tempDate = changeDateSeparatorToSlash(tempDate)
    
    return tempDate

def monthAlphaToNum(tempDate):

    if "jan" in tempDate:
        tempDate = tempDate.replace("jan", "01")
    elif "feb" in tempDate:
        tempDate = tempDate.replace("feb", "02")
    elif "mar" in tempDate:
        tempDate = tempDate.replace("mar", "03")
    elif "apr" in tempDate:
        tempDate = tempDate.replace("apr", "04")
    elif "may" in tempDate:
        tempDate = tempDate.replace("may", "05")
    elif "jun" in tempDate:
        tempDate = tempDate.replace("jun", "06")
    elif "jul" in tempDate:
        tempDate = tempDate.replace("jul", "07")
    elif "aug" in tempDate:
        tempDate = tempDate.replace("aug", "08")
    elif "sep" in tempDate:
        tempDate = tempDate.replace("sep", "09")
    elif "oct" in tempDate:
        tempDate = tempDate.replace("oct", "10")
    elif "nov" in tempDate:
        tempDate = tempDate.replace("nov", "11")
    elif "dec" in tempDate:
        tempDate = tempDate.replace("dec", "12")
    return tempDate

def cleanDateTime_YearAtStart(tempDate):
   
    tempDate = changeDateSeparatorToSlash(monthAlphaToNum(tempDate))

    if re.search(r'^(\d{2}/\d+/\d+ )', tempDate):
        splitted_tempDate = tempDate.split("/")
        year = splitted_tempDate[0]

        if int(year) > 40:
            tempDate = "19" + year + "/" + splitted_tempDate[1] + "/" + splitted_tempDate[2]
            
        else:
            tempDate = "20" + year + "/" + splitted_tempDate[1] + "/" + splitted_tempDate[2]

    if len(tempDate.split(" ")[0])  == 9:
        if len(splitted_tempDate[1]) == 1:
            tempDate = tempDate[:5] + "0" + tempDate[5:]

        elif len(splitted_tempDate[2]) == 1:
            tempDate = tempDate[:8] + "0" + tempDate[8:]

    elif len(tempDate.split(" ")[0]) == 8:
            tempDate = tempDate[:5] + "0" + tempDate[5:]
            tempDate = tempDate[:8] + "0" + tempDate[8:]
 
    return tempDate


def cleanDateTime_YearAtEnd(tempDate):
    
    tempDate = changeDateSeparatorToSlash(monthAlphaToNum(tempDate))

    splitted_tempDate = tempDate.split("/")
    if re.search(r'^(\d+/\d+/\d{2} )', tempDate):
        year = splitted_tempDate[2][:2]

        if int(year) > 40:
            tempDate = splitted_tempDate[0] + "/" + splitted_tempDate[1] + "/" + "19" + splitted_tempDate[2]
            
        else:
            tempDate = splitted_tempDate[0] + "/" + splitted_tempDate[1] + "/" + "20" + splitted_tempDate[2]

    if len(tempDate.split(" ")[0])  == 9:
        if len(splitted_tempDate[0]) == 1:
            tempDate = "0" + tempDate

        elif len(splitted_tempDate[1]) == 1:
            tempDate = tempDate[:3] + "0" + tempDate[3:]

    elif len(tempDate.split(" ")[0]) == 8:
            tempDate = "0" + tempDate
            tempDate = tempDate[:3] + "0" + tempDate[3:]

    return tempDate


def cleanDate_YearAtStart(tempDate):
    
    tempDate = changeDateSeparatorToSlash(monthAlphaToNum(tempDate))
    
    splitted_tempDate = tempDate.split("/")
    if re.search(r'^(\d{2}/\d+/\d+)$', tempDate):
        splitted_tempDate = tempDate.split("/")
        year = splitted_tempDate[0]

        if int(year) > 40:
            tempDate = "19" + year + "/" + splitted_tempDate[1] + "/" + splitted_tempDate[2]
            
        else:
            tempDate = "20" + year + "/" + splitted_tempDate[1] + "/" + splitted_tempDate[2]

    if len(splitted_tempDate)  == 9:
        if len(splitted_tempDate[1]) == 1:
            tempDate = tempDate[:5] + "0" + tempDate[5:]

        elif len(splitted_tempDate[2]) == 1:
            tempDate = tempDate[:8] + "0" + tempDate[8:]

    elif len(splitted_tempDate) == 8:
            tempDate = tempDate[:5] + "0" + tempDate[5:]
            tempDate = tempDate[:8] + "0" + tempDate[8:]
            
    return tempDate


def cleanDate_YearAtEnd(tempDate):
    
    tempDate = changeDateSeparatorToSlash(monthAlphaToNum(tempDate))

    splitted_tempDate = tempDate.split("/")
    if re.search(r'^(\d+/\d+/\d{2})$', tempDate):
        year = splitted_tempDate[2]

        if int(year) > 40:
            tempDate = splitted_tempDate[0] + "/" + splitted_tempDate[1] + "/" + "19" + splitted_tempDate[2]
            
        else:
            tempDate = splitted_tempDate[0] + "/" + splitted_tempDate[1] + "/" + "20" + splitted_tempDate[2]

    if len(tempDate)  == 9:
        if len(splitted_tempDate[0]) == 1:
            tempDate = "0" + tempDate

        elif len(splitted_tempDate[1]) == 1:
            tempDate = tempDate[:3] + "0" + tempDate[3:]

    elif len(tempDate) == 8:
            tempDate = "0" + tempDate
            tempDate = tempDate[:3] + "0" + tempDate[3:]   

    return tempDate


def cleanData(raw_data_list, header_type, sub_type, format_type, chop_zero):
    len_single_row = len(header_type)
    len_raw_data_list = len(raw_data_list)
    chopped_indices = []
    for i in range(len_raw_data_list):
        for j in range(len_single_row):
            if header_type[j] == datatypes.HEADER_TYPES['string']:
                if not raw_data_list[i][j].replace(" ", "") == "":
                    pass
                else:
                    raw_data_list[i][j] = ""

            elif header_type[j] == datatypes.HEADER_TYPES['number']:
                if sub_type[j] == datatypes.NUMBER_SUBTYPES['percentage']:
                    raw_data_list[i][j] = cleanNumberPercentage(raw_data_list[i][j])
                elif sub_type[j] == datatypes.NUMBER_SUBTYPES['currency']:
                    raw_data_list[i][j] = cleanNumberCurrency(raw_data_list[i][j])
                elif sub_type[j] == datatypes.NUMBER_SUBTYPES['normal']:
                    raw_data_list[i][j] = cleanNumberNumber(raw_data_list[i][j])
                else:
                    pass

            else:
                if sub_type[j] == datatypes.DATE_SUBTYPES['dateonly']:
                    if re.search(r'(YYYY)$', format_type[j]):
                        if not raw_data_list[i][j].replace(" ", "") == "":
                            
                            raw_data_list[i][j] = cleanDate_YearAtEnd(raw_data_list[i][j])
                            
                            x, y, year = raw_data_list[i][j].split("/")
                            if not isLeapYear(int(year)):
                                if (int(x) + int(y)) == 31 and (x == "02" or y == "02"):
                                    assignTypes(format_type, sub_type, header_type, j, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                            
                    elif re.search(r'^(YYYY)', format_type[j]):
                        
                        if not raw_data_list[i][j] == "":
                             
                            raw_data_list[i][j] = cleanDate_YearAtStart(raw_data_list[i][j])
                            
                            year, x, y = raw_data_list[i][j].split("/")
                            if not isLeapYear(int(year)):
                                if (int(x) + int(y)) == 31 and (x == "02" or y == "02"):
                                    assignTypes(format_type, sub_type, header_type, j, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                            
                elif sub_type[j] == datatypes.DATE_SUBTYPES['datetime']:
                    if chop_zero[j] == 1:
                        if re.search(r'(YYYY) ', format_type[j]):
                            if not raw_data_list[i][j] == "":
                                raw_data_list[i][j] = raw_data_list[i][j].replace(" 00:00:00", "")
                                raw_data_list[i][j] = cleanDate_YearAtEnd(raw_data_list[i][j])
                                
                        elif re.search(r'^(YYYY)', format_type[j]):
                            
                            if not raw_data_list[i][j] == "":
                                raw_data_list[i][j] = raw_data_list[i][j].replace(" 00:00:00", "")
                                raw_data_list[i][j] = cleanDate_YearAtStart(raw_data_list[i][j])
                                
                        chopped_indices.append(j)
                                
                    else:
                        if re.search(r'(YYYY) ', format_type[j]):
                            if not raw_data_list[i][j] == "":
                                
                                raw_data_list[i][j] = cleanDateTime_YearAtEnd(raw_data_list[i][j])
                                
                                x, y, year = raw_data_list[i][j].split(" ")[0].split("/")
                                if not isLeapYear(int(year)):
                                    if (int(x) + int(y)) == 31 and (x == "02" or y == "02"):
                                        assignTypes(format_type, sub_type, header_type, j, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                        
                        elif re.search(r'^(YYYY)', format_type[j]):
                            
                            if not raw_data_list[i][j] == "":
                                
                                raw_data_list[i][j] = cleanDateTime_YearAtStart(raw_data_list[i][j])
                                
                                year, x, y = raw_data_list[i][j].split(" ")[0].split("/")
                                if not isLeapYear(int(year)):
                                    if (int(x) + int(y)) == 31 and (x == "02" or y == "02"):
                                        assignTypes(format_type, sub_type, header_type, j, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                
                else:
                    pass
                
    
    for i in range(len_single_row):
        if header_type[i] == datatypes.HEADER_TYPES['_']:
            assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
            
    if len(chopped_indices) > 0:
        for i in chopped_indices:
            sub_type[i] = sub_type[i].replace("TIME", "")
            format_type[i] = format_type[i].replace(" HH:mm:ss", "")

    return [raw_data_list, header_type, sub_type, format_type]

######### NEW METHODS END ########################################

def isAgainstHeaderIndices(lst, header_type, sub_type, format_type, headers_list, deletedHeaders_asListRev, first_data_index, chop_zero, user_correction_status):#####Include the added arguments where the method is called
    #type identification
    temp_bool = False
    i = 0
    single_row = []
#    againstHeader_checkRequired = True
    
    if user_correction_status == config.USER_CORRECTION_STATUS['success']:
        lst = neglect_indices(lst, deletedHeaders_asListRev)
    
    while i < len(lst):
        element = lst[i]
        tempDate = element.strip(" ").lower()
        temp = element.replace(" ", "").lower()
        
#        if againstHeader_checkRequired:
#            if (not isCellEmpty(temp)):
#                if i < first_nonemptyIndex:
#                    temp_bool = True
#                    break
#                else:
#                    againstHeader_checkRequired = False
            
        #_________________________________________________________________
        if (isCellEmpty(temp)) and header_type[i] == datatypes.HEADER_TYPES['_']:
            single_row.append("")
        #_________________________________________________________________

        else:
            if header_type[i] == datatypes.HEADER_TYPES['number']:
                if re.search(regex.NUMBER_REGEX, temp):
                    ################################
                        #START(1)(ii)
                    ################################
                    if sub_type[i] == datatypes.NUMBER_SUBTYPES['currency']:
                        ################################
                                #START(1)(ii)(a)
                        ################################
                        if isExistsSpecificCurrencySymbolInNumber(temp):
                            
                            #format_type[i] == "format not updated": case is left out here because format_type list will have been updated by now
                            if format_type[i] == datatypes.NUMBER_FORMAT_TYPES['currency1']:
                                if datatypes.NUMBER_FORMAT_TYPES['currency1'] in temp:
                                    pass
                                else:
                                    assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                    
                            elif format_type[i] == datatypes.NUMBER_FORMAT_TYPES['currency2']:
                                if datatypes.NUMBER_FORMAT_TYPES['currency2'] in temp:
                                    pass
                                else:
                                    assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                            
                            elif format_type[i] == datatypes.NUMBER_FORMAT_TYPES['currency3']:
                                if datatypes.NUMBER_FORMAT_TYPES['currency3'] in temp:
                                    pass
                                else:
                                    assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                    
                            elif format_type[i] == datatypes.NUMBER_FORMAT_TYPES['currency4']:
                                if datatypes.NUMBER_FORMAT_TYPES['currency4'] in temp:
                                    pass
                                else:
                                    assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                    
                            else:
                                pass
                        ################################
                                #END(1)(ii)(a)
                        ################################
                                
                        ################################
                                #START(1)(ii)(b)
                        ################################
                        else:
                            assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                        ################################
                                #END(1)(ii)(b)
                        ################################
                        
                    ################################
                        #END(1)(ii)
                    ################################        
                    
                    ################################
                        #START(1)(iii)
                    ################################
                    elif sub_type[i] == datatypes.NUMBER_SUBTYPES['percentage']:
                        ################################
                                #START(1)(iii)(a)
                        ################################
                        if datatypes.NUMBER_FORMAT_TYPES['percentage'] in temp:
                            pass
                        ################################
                                #END(1)(iii)(a)
                        ################################
                        
                        ################################
                                #START(1)(iii)(b)
                        ################################
                        else:
                            assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                        ################################
                                #END(1)(iii)(b)
                        ################################
                        
                    ################################
                        #END(1)(iii)
                    ################################
                    
                    ################################
                        #START(1)(iv)
                    ################################
                    elif sub_type[i] == datatypes.NUMBER_SUBTYPES['normal']:
                        for symbol in datatypes.CURRENCY_SYMBOLS + ["%"]:
                            if symbol in temp:
                                assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                break
                    ################################
                        #END(1)(iv)
                    ################################
                #_________________________________________________________________
                elif isCellEmpty(temp):
                    pass
                #_________________________________________________________________
                else:
                    assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
            ###################################################################
            ###################################################################
            
            ###################################################################
            ###################################################################
            elif header_type[i] == datatypes.HEADER_TYPES['date']:
                if re.search(regex.DATE_REGEX, tempDate):
                    ################################
                        #START(2)(ii)
                    ################################
                    if sub_type[i] == datatypes.DATE_SUBTYPES['dateonly']:
                        ################################
                                #START(2)(ii)(a)
                        ################################
                        if re.search(regex.DATEONLY_REGEX, tempDate):
                            
                            if format_type[i] == datatypes.DATEONLY_FORMAT_TYPES['YDM']:
                                if re.search(regex.DATEONLY_YDM_REGEX, tempDate):
                                    pass
                                else:
                                    assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                    
                            elif format_type[i] == datatypes.DATEONLY_FORMAT_TYPES['YMD']:
                                if re.search(regex.DATEONLY_YMD_REGEX, tempDate):
                                    pass
                                
                                elif re.search(regex.DATEONLY_YDM_REGEX, tempDate):
                                    assignTypes(format_type, sub_type, header_type, i, datatypes.DATEONLY_FORMAT_TYPES['YDM'], datatypes.DATE_SUBTYPES['dateonly'], datatypes.HEADER_TYPES['date'])
                                    
                                else:
                                    assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                    
                            elif format_type[i] == datatypes.DATEONLY_FORMAT_TYPES['DMY']:
                                if re.search(regex.DATEONLY_DMY_REGEX, tempDate):
                                    pass
                                
                                elif re.search(regex.DATEONLY_YMD_REGEX, tempDate):
                                    assignTypes(format_type, sub_type, header_type, i, datatypes.DATEONLY_FORMAT_TYPES['YMD'], datatypes.DATE_SUBTYPES['dateonly'], datatypes.HEADER_TYPES['date'])
                                    
                                elif re.search(regex.DATEONLY_YDM_REGEX, tempDate):
                                    assignTypes(format_type, sub_type, header_type, i, datatypes.DATEONLY_FORMAT_TYPES['YDM'], datatypes.DATE_SUBTYPES['dateonly'], datatypes.HEADER_TYPES['date'])
                                    
                                else:
                                    assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                            
                            elif format_type[i] == datatypes.DATEONLY_FORMAT_TYPES['MDY']:
                                if re.search(regex.DATEONLY_MDY_REGEX, tempDate):
                                    pass
                                
                                elif re.search(regex.DATEONLY_DMY_REGEX, tempDate):
                                    assignTypes(format_type, sub_type, header_type, i, datatypes.DATEONLY_FORMAT_TYPES['DMY'], datatypes.DATE_SUBTYPES['dateonly'], datatypes.HEADER_TYPES['date'])
                                    
                                elif re.search(regex.DATEONLY_YMD_REGEX, tempDate):
                                    assignTypes(format_type, sub_type, header_type, i, datatypes.DATEONLY_FORMAT_TYPES['YMD'], datatypes.DATE_SUBTYPES['dateonly'], datatypes.HEADER_TYPES['date'])
                                    
                                elif re.search(regex.DATEONLY_YDM_REGEX, tempDate):
                                    assignTypes(format_type, sub_type, header_type, i, datatypes.DATEONLY_FORMAT_TYPES['YDM'], datatypes.DATE_SUBTYPES['dateonly'], datatypes.HEADER_TYPES['date'])
                                    
                                else:
                                    assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                            
                            else:
                                pass
                        ################################
                                #END(2)(ii)(a)
                        ################################
                        #_________________________________________________________________
                        elif isCellEmpty(temp):
                            pass
                        #_________________________________________________________________
                        
                        ################################
                                #START(2)(ii)(b)
                        ################################
                        else:
                            assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                        ################################
                                #END(2)(ii)(b)
                        ################################
                    ################################
                        #END(2)(ii)
                    ################################
                    
                    ################################
                        #START(2)(iii)
                    ################################
                    elif sub_type[i] == datatypes.DATE_SUBTYPES['datetime']:
                        ################################
                                #START(2)(iii)(a)
                        ################################
                        if re.search(regex.DATETIME_REGEX, tempDate):
                                        
                            if re.search(r'^YYYY/DD', format_type[i]):
                                
                                if format_type[i] == datatypes.DATETIME_FORMAT_TYPES['YDM1']:
                                    if re.search(regex.DATETIME_YDM1_REGEX, tempDate):
                                        pass
                                    else:
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                
                                elif format_type[i] == datatypes.DATETIME_FORMAT_TYPES['YDM2']:
                                    if re.search(regex.DATETIME_YDM2_REGEX, tempDate):
                                        pass
                                    else:
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                
                                elif format_type[i] == datatypes.DATETIME_FORMAT_TYPES['YDM3']:
                                    if re.search(regex.DATETIME_YDM3_REGEX, tempDate):
                                        pass
                                    else:
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                        
                                elif format_type[i] == datatypes.DATETIME_FORMAT_TYPES['YDM4']:
                                    if re.search(regex.DATETIME_YDM4_REGEX, tempDate):
                                        pass
                                    else:
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                    
                                elif format_type[i] == datatypes.DATETIME_FORMAT_TYPES['YDM5']:
                                    if re.search(regex.DATETIME_YDM5_REGEX, tempDate):
                                        pass
                                    else:
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                    
                            elif re.search(r'^YYYY/MM', format_type[i]):
                                
                                if format_type[i] == datatypes.DATETIME_FORMAT_TYPES['YMD1']:
                                    if re.search(regex.DATETIME_YMD1_REGEX, tempDate):
                                        pass
                                    
                                    elif re.search(regex.DATETIME_YDM1_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YDM1'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                    
                                    else:
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                
                                elif format_type[i] == datatypes.DATETIME_FORMAT_TYPES['YMD2']:
                                    if re.search(regex.DATETIME_YMD2_REGEX, tempDate):
                                        pass
                                    
                                    elif re.search(regex.DATETIME_YDM2_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YDM2'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                        
                                    else:
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                
                                elif format_type[i] == datatypes.DATETIME_FORMAT_TYPES['YMD3']:
                                    if re.search(regex.DATETIME_YMD3_REGEX, tempDate):
                                        pass
                                    
                                    elif re.search(regex.DATETIME_YDM3_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YDM3'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                        
                                    else:
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                        
                                elif format_type[i] == datatypes.DATETIME_FORMAT_TYPES['YMD4']:
                                    if re.search(regex.DATETIME_YMD4_REGEX, tempDate):
                                        pass
                                    
                                    elif re.search(regex.DATETIME_YDM4_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YDM4'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                        
                                    else:
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                    
                                elif format_type[i] == datatypes.DATETIME_FORMAT_TYPES['YMD5']:
                                    if re.search(regex.DATETIME_YMD5_REGEX, tempDate):
                                        pass
                                    
                                    elif re.search(regex.DATETIME_YDM5_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YDM5'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                        
                                    else:
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                    
                            elif re.search(r'^DD/MM', format_type[i]):
                                
                                if format_type[i] == datatypes.DATETIME_FORMAT_TYPES['DMY1']:
                                    if re.search(regex.DATETIME_DMY1_REGEX, tempDate):
                                        pass
                                    
                                    elif re.search(regex.DATETIME_YMD1_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YMD1'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                    
                                    elif re.search(regex.DATETIME_YDM1_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YDM1'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                    
                                    else:
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                
                                elif format_type[i] == datatypes.DATETIME_FORMAT_TYPES['DMY2']:
                                    if re.search(regex.DATETIME_DMY2_REGEX, tempDate):
                                        pass
                                    
                                    elif re.search(regex.DATETIME_YMD2_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YMD2'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                    
                                    elif re.search(regex.DATETIME_YDM2_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YDM2'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                    
                                    else:
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                
                                elif format_type[i] == datatypes.DATETIME_FORMAT_TYPES['DMY3']:
                                    if re.search(regex.DATETIME_DMY3_REGEX, tempDate):
                                        pass
                                    
                                    elif re.search(regex.DATETIME_YMD3_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YMD3'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                    
                                    elif re.search(regex.DATETIME_YDM3_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YDM3'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                    
                                    else:
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                        
                                elif format_type[i] == datatypes.DATETIME_FORMAT_TYPES['DMY4']:
                                    if re.search(regex.DATETIME_DMY4_REGEX, tempDate):
                                        pass
                                    
                                    elif re.search(regex.DATETIME_YMD4_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YMD4'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                    
                                    elif re.search(regex.DATETIME_YDM4_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YDM4'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                    
                                    else:
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                    
                                elif format_type[i] == datatypes.DATETIME_FORMAT_TYPES['DMY5']:
                                    if re.search(regex.DATETIME_DMY5_REGEX, tempDate):
                                        pass
                                    
                                    elif re.search(regex.DATETIME_YMD5_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YMD5'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                     
                                    elif re.search(regex.DATETIME_YDM5_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YDM5'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                    
                                    else:
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                    
                            elif re.search(r'^MM/DD', format_type[i]):
                                
                                if format_type[i] == datatypes.DATETIME_FORMAT_TYPES['MDY1']:
                                    if re.search(regex.DATETIME_MDY1_REGEX, tempDate):
                                        pass
                                    
                                    elif re.search(regex.DATETIME_DMY1_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['DMY1'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                    
                                    elif re.search(regex.DATETIME_YMD1_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YMD1'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                    
                                    elif re.search(regex.DATETIME_YDM1_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YDM1'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                    
                                    else:
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                    
                                elif format_type[i] == datatypes.DATETIME_FORMAT_TYPES['MDY2']:
                                    if re.search(regex.DATETIME_MDY2_REGEX, tempDate):
                                        pass
                                    
                                    elif re.search(regex.DATETIME_DMY2_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['DMY2'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                    
                                    elif re.search(regex.DATETIME_YMD2_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YMD2'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                    
                                    elif re.search(regex.DATETIME_YDM2_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YDM2'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                    
                                    else:
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                
                                elif format_type[i] == datatypes.DATETIME_FORMAT_TYPES['MDY3']:
                                    if re.search(regex.DATETIME_MDY3_REGEX, tempDate):
                                        pass
                                    
                                    elif re.search(regex.DATETIME_DMY3_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['DMY3'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                    
                                    elif re.search(regex.DATETIME_YMD3_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YMD3'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                    
                                    elif re.search(regex.DATETIME_YDM3_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YDM3'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                    
                                    else:
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                        
                                elif format_type[i] == datatypes.DATETIME_FORMAT_TYPES['MDY4']:
                                    if re.search(regex.DATETIME_MDY4_REGEX, tempDate):
                                        pass
                                    
                                    elif re.search(regex.DATETIME_DMY4_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['DMY4'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                    
                                    elif re.search(regex.DATETIME_YMD4_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YMD4'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                    
                                    elif re.search(regex.DATETIME_YDM4_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YDM4'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                    
                                    else:
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                    
                                elif format_type[i] == datatypes.DATETIME_FORMAT_TYPES['MDY5']:
                                    if re.search(regex.DATETIME_MDY5_REGEX, tempDate):
                                        pass
                                    
                                    elif re.search(regex.DATETIME_DMY5_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['DMY5'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                    
                                    elif re.search(regex.DATETIME_YMD5_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YMD5'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                    
                                    elif re.search(regex.DATETIME_YDM5_REGEX, tempDate):
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YDM5'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                    
                                    else:
                                        assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                    
                            else:
                                pass
                        ################################
                                #END(2)(iii)(a)
                        ################################
                        ################################
                                #START(2)(iii)(b)
                        ################################
                        
                        else:
                            assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                        
                        ################################
                                #END(2)(iii)(b)
                        ################################
                    ################################
                        #END(2)(iii)
                    ################################
                    
                    ################################
                        #START(2)(iv)
                    ################################
                    
                    elif sub_type[i] == datatypes.DATE_SUBTYPES['time']:
                        ################################
                                #START(2)(iv)(a)
                        ################################
                        if re.search(regex.TIME_REGEX, temp):
                        
                            if format_type[i] == datatypes.TIME_FORMAT_TYPES['time1']:
                                
                                if re.search(regex.TIME1_REGEX, temp):
                                    pass
                                else:
                                    assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                
                            elif format_type[i] == datatypes.TIME_FORMAT_TYPES['time2']:
                                
                                if re.search(regex.TIME2_REGEX, temp):
                                    pass
                                else:
                                    assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                
                            elif format_type[i] == datatypes.TIME_FORMAT_TYPES['time3']:
                                
                                if re.search(regex.TIME3_REGEX, temp):
                                    pass
                                else:
                                    assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                    
                            else:
                                pass
                                
                            
                        
                        elif re.search(regex.DURATION_OR_TIME_REGEX, temp) and not re.search(r'duration', headers_list[i - first_data_index].lower()):
                        #Ensuring that temp is not categorized under 'duration'...
                            
                            if format_type[i] == datatypes.TIME_FORMAT_TYPES['time4']:
                                
                                if re.search(regex.TIME4_REGEX, temp):
                                    pass
                                else:
                                    assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                
                            elif format_type[i] == datatypes.TIME_FORMAT_TYPES['time5']:
                                
                                if re.search(regex.TIME5_REGEX,temp):
                                    pass
                                else:
                                    assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                    
                            else:
                                pass
                        
                        ################################
                                #END(2)(iii)(b)
                        ################################
                        ################################
                                #START(2)(iv)(b)
                        ################################
                        
                        else:
                            assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                        
                        ################################
                                #END(2)(iv)(b)
                        ################################
                    ################################
                        #END(2)(iv)
                    ################################
                    
                    ################################
                        #START(2)(v)
                    ################################
                    elif sub_type[i] == datatypes.DATE_SUBTYPES['duration']:
                        ################################
                                #START(2)(v)(a)
                        ################################
                        
                        if re.search(regex.DURATION_REGEX, temp):
                                
                            if format_type[i] == datatypes.TIME_FORMAT_TYPES['duration1']:
                                
                                if re.search(regex.DURATION1_REGEX, temp):
                                    pass
                                else:
                                    assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                            
                            elif format_type[i] == datatypes.TIME_FORMAT_TYPES['duration2']:
                                
                                if re.search(regex.DURATION2_REGEX, temp):
                                    pass
                                else:
                                    assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                    
                            elif format_type[i] == datatypes.TIME_FORMAT_TYPES['duration3']:
                                
                                if re.search(regex.DURATION3_REGEX, temp):
                                    pass
                                else:
                                    assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                                    
                            else:
                                pass
                        
                        ################################
                                #END(2)(v)(a)
                        ################################
                        ################################
                                #START(2)(v)(b)
                        ################################
                        
                        else:
                            assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                        
                        ################################
                                #END(2)(v)(b)
                        ################################
                    ################################
                        #END(2)(v)
                    ################################
                    ################################
                        #START(2)(vi)
                    ################################
                    
                    else:
                        pass
                    
                    ################################
                        #END(2)(vi)
                    ################################
                #_________________________________________________________________
                elif isCellEmpty(temp):
                    pass
                #_________________________________________________________________
                else:
                    assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
            ###################################################################
            ###################################################################
            
            elif header_type[i] == datatypes.HEADER_TYPES['string']:
                pass
            
            else:
                ################################
                        #START(1)
                ################################
                if re.search(regex.NUMBER_REGEX, temp):
                    
                    ################################
                        #START(1)(i)
                    ################################
#                    if sub_type[i] == "sub not updated":
                    ################################
                            #START(1)(i)(a)
                    ################################
                    if isExistsSpecificCurrencySymbolInNumber(temp):
                    
                        if datatypes.NUMBER_FORMAT_TYPES['currency1'] in temp:
                            assignTypes(format_type, sub_type, header_type, i, datatypes.NUMBER_FORMAT_TYPES['currency1'], datatypes.NUMBER_SUBTYPES['currency'], datatypes.HEADER_TYPES['number'])
                        elif datatypes.NUMBER_FORMAT_TYPES['currency2'] in temp:
                            assignTypes(format_type, sub_type, header_type, i, datatypes.NUMBER_FORMAT_TYPES['currency2'], datatypes.NUMBER_SUBTYPES['currency'], datatypes.HEADER_TYPES['number'])
                        elif datatypes.NUMBER_FORMAT_TYPES['currency3'] in temp:
                            assignTypes(format_type, sub_type, header_type, i, datatypes.NUMBER_FORMAT_TYPES['currency3'], datatypes.NUMBER_SUBTYPES['currency'], datatypes.HEADER_TYPES['number'])
                        elif datatypes.NUMBER_FORMAT_TYPES['currency4'] in temp:
                            assignTypes(format_type, sub_type, header_type, i, datatypes.NUMBER_FORMAT_TYPES['currency4'], datatypes.NUMBER_SUBTYPES['currency'], datatypes.HEADER_TYPES['number'])
                    ################################
                            #END(1)(i)(a)
                    ################################
                            
                    ################################
                            #START(1)(i)(b)
                    ################################
                    elif datatypes.NUMBER_FORMAT_TYPES['percentage'] in temp:
                        
                        assignTypes(format_type, sub_type, header_type, i, datatypes.NUMBER_FORMAT_TYPES['percentage'], datatypes.NUMBER_SUBTYPES['percentage'], datatypes.HEADER_TYPES['number'])
                    ################################
                            #END(1)(i)(b)
                    ################################
                    
                    ################################
                            #START(1)(i)(c)
                    ################################
                    else:#if not percentage or currency, then should definitely be just a number
                        assignTypes(format_type, sub_type, header_type, i, datatypes.NUMBER_FORMAT_TYPES['nothing'], datatypes.NUMBER_SUBTYPES['normal'], datatypes.HEADER_TYPES['number'])
                    ################################
                            #END(1)(i)(c)
                    ################################
                ################################
                        #END(1)
                ################################
                
                ################################
                        #START(2)
                ################################    
                elif re.search(regex.DATE_REGEX, tempDate):
                    
                    ################################
                        #START(2)(i)
                    ################################
#                    if sub_type[i] == "sub not updated":
                    ################################
                            #START(2)(i)(a) date only start
                    ################################
                    if re.search(regex.DATEONLY_REGEX, tempDate):
                        #N.B: assignTypes method is not executed under this condition on purpose
                        if re.search(regex.DATEONLY_YDM_REGEX, tempDate):
                            #YDM
                            format_type[i] = datatypes.DATEONLY_FORMAT_TYPES['YDM']
                            
                        if re.search(regex.DATEONLY_YMD_REGEX, tempDate):
                            #YMD
                            format_type[i] = datatypes.DATEONLY_FORMAT_TYPES['YMD']
                        
                        if re.search(regex.DATEONLY_DMY_REGEX, tempDate):
                            #DMY
                            format_type[i] = datatypes.DATEONLY_FORMAT_TYPES['DMY']
                            
                        if re.search(regex.DATEONLY_MDY_REGEX, tempDate):
                            #MDY
                            format_type[i] = datatypes.DATEONLY_FORMAT_TYPES['MDY']
                            
                        sub_type[i] = datatypes.DATE_SUBTYPES['dateonly']
                        header_type[i] = datatypes.HEADER_TYPES['date']
                    ################################
                            #END(2)(i)(a) date only end
                    ################################
                    
                    ################################
                            #START(2)(i)(b) DATETIME start
                    ################################
                    elif re.search(regex.DATETIME_REGEX, tempDate):
                        
                        if re.search(regex.DATETIME_YDM_PREFIX_REGEX, tempDate):
                            #YDM
                            if re.search(regex.DATETIME_TIME1_SUFFIX_REGEX, tempDate):
                                #YDM HH:mm A NB: if HH = [1-9], then add a '0' before it ere appending to data_list
                                assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YDM1'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                
                            elif re.search(regex.DATETIME_TIME2_SUFFIX_REGEX, tempDate):
                                #YDM HH:mm:ss
                                assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YDM2'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                
                            elif re.search(regex.DATETIME_TIME3_SUFFIX_REGEX, tempDate):
                                #YDM H:mm:ss
                                assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YDM3'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                
                            elif re.search(regex.DATETIME_TIME4_SUFFIX_REGEX, tempDate):
                                #YDM HH:mm
                                assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YDM4'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                
                            elif re.search(regex.DATETIME_TIME5_SUFFIX_REGEX, tempDate):
                                #YDM H:mm
                                assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YDM5'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                
                        if re.search(regex.DATETIME_YMD_PREFIX_REGEX, tempDate):
                            #YMD
                            if re.search(regex.DATETIME_TIME1_SUFFIX_REGEX, tempDate):
                                #YMD HH:mm A
                                assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YMD1'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                
                            elif re.search(regex.DATETIME_TIME2_SUFFIX_REGEX, tempDate):
                                #YMD HH:mm:ss
                                assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YMD2'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                
                            elif re.search(regex.DATETIME_TIME3_SUFFIX_REGEX, tempDate):
                                #YMD H:mm:ss
                                assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YMD3'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                
                            elif re.search(regex.DATETIME_TIME4_SUFFIX_REGEX, tempDate):
                                #YMD HH:mm
                                assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YMD4'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                
                            elif re.search(regex.DATETIME_TIME5_SUFFIX_REGEX, tempDate):
                                #YMD H:mm
                                assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['YMD5'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                
                        if re.search(regex.DATETIME_DMY_PREFIX_REGEX, tempDate):
                            #DMY
                            if re.search(regex.DATETIME_TIME1_SUFFIX_REGEX, tempDate):
                                #DMY HH:mm A
                                assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['DMY1'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                
                            elif re.search(regex.DATETIME_TIME2_SUFFIX_REGEX, tempDate):
                                #DMY HH:mm:ss
                                assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['DMY2'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                
                            elif re.search(regex.DATETIME_TIME3_SUFFIX_REGEX, tempDate):
                                #DMY H:mm:ss
                                assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['DMY3'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                
                            elif re.search(regex.DATETIME_TIME4_SUFFIX_REGEX, tempDate):
                                #DMY HH:mm
                                assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['DMY4'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                
                            elif re.search(regex.DATETIME_TIME5_SUFFIX_REGEX, tempDate):
                                #DMY H:mm
                                assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['DMY5'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                            
                        if re.search(regex.DATETIME_MDY_PREFIX_REGEX, tempDate):
                            #MDY
                            if re.search(regex.DATETIME_TIME1_SUFFIX_REGEX, tempDate):
                                #MDY HH:mm A
                                assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['MDY1'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                
                            elif re.search(regex.DATETIME_TIME2_SUFFIX_REGEX, tempDate):
                                #MDY HH:mm:ss
                                assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['MDY2'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                
                            elif re.search(regex.DATETIME_TIME3_SUFFIX_REGEX, tempDate):
                                #MDY H:mm:ss
                                assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['MDY3'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                
                            elif re.search(regex.DATETIME_TIME4_SUFFIX_REGEX, tempDate):
                                #MDY HH:mm
                                assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['MDY4'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                                
                            elif re.search(regex.DATETIME_TIME5_SUFFIX_REGEX, tempDate):
                                #MDY H:mm
                                assignTypes(format_type, sub_type, header_type, i, datatypes.DATETIME_FORMAT_TYPES['MDY5'], datatypes.DATE_SUBTYPES['datetime'], datatypes.HEADER_TYPES['date'])
                    ################################
                            #END(2)(i)(b) DATETIME end
                    ################################
                    
                    ################################
                            #START(2)(i)(c) TIME start
                    ################################
                    
                    elif re.search(regex.TIME_REGEX, tempDate):
                        
                        if re.search(regex.TIME1_REGEX, tempDate):
                            # HH:mm A
                            assignTypes(format_type, sub_type, header_type, i, datatypes.TIME_FORMAT_TYPES['time1'], datatypes.DATE_SUBTYPES['time'], datatypes.HEADER_TYPES['date'])
                            
                        elif re.search(regex.TIME2_REGEX, tempDate):
                            # HH:mm:ss
                            assignTypes(format_type, sub_type, header_type, i, datatypes.TIME_FORMAT_TYPES['time2'], datatypes.DATE_SUBTYPES['time'], datatypes.HEADER_TYPES['date'])
                            
                        elif re.search(regex.TIME3_REGEX, tempDate):
                            # H:mm:ss
                            assignTypes(format_type, sub_type, header_type, i, datatypes.TIME_FORMAT_TYPES['time3'], datatypes.DATE_SUBTYPES['time'], datatypes.HEADER_TYPES['date'])
                    
                    elif re.search(regex.DURATION_OR_TIME_REGEX, tempDate) and not re.search(r'duration', headers_list[i - first_data_index].lower()):
                    #Ensuring that temp is not categorized under 'duration'...
                        
                        if re.search(regex.TIME4_REGEX, tempDate):
                            # HH:mm
                            assignTypes(format_type, sub_type, header_type, i, datatypes.TIME_FORMAT_TYPES['time4'], datatypes.DATE_SUBTYPES['time'], datatypes.HEADER_TYPES['date'])
                            
                        elif re.search(regex.TIME5_REGEX, tempDate):
                            # H:mm
                            assignTypes(format_type, sub_type, header_type, i, datatypes.TIME_FORMAT_TYPES['time5'], datatypes.DATE_SUBTYPES['time'], datatypes.HEADER_TYPES['date'])
                            
                    ################################
                            #END(2)(i)(c) TIME end
                    ################################
                    
                    ################################
                            #START(2)(i)(d) DURATION start
                    ################################
                    
                    elif re.search(regex.DURATION_REGEX, tempDate):
                    
                        if re.search(regex.DURATION1_REGEX, tempDate):
                            assignTypes(format_type, sub_type, header_type, i, datatypes.TIME_FORMAT_TYPES['duration1'], datatypes.DATE_SUBTYPES['duration'], datatypes.HEADER_TYPES['date'])
                            
                        elif re.search(regex.DURATION2_REGEX, tempDate):
                            assignTypes(format_type, sub_type, header_type, i, datatypes.TIME_FORMAT_TYPES['duration2'], datatypes.DATE_SUBTYPES['duration'], datatypes.HEADER_TYPES['date'])
                            
                        elif re.search(regex.DURATION3_REGEX, tempDate):
                            assignTypes(format_type, sub_type, header_type, i, datatypes.TIME_FORMAT_TYPES['duration3'], datatypes.DATE_SUBTYPES['duration'], datatypes.HEADER_TYPES['date'])
                        
                    ################################
                            #END(2)(i)(d) DURATION end
                    ################################
                    ################################
                        #END(2)(i)
                    ################################
                    

                ################################
                        #END(2)
                ################################
                
                ################################
                        #START(3)
                ################################
                else:
                    assignTypes(format_type, sub_type, header_type, i, datatypes.STRING_FORMAT_TYPES['nothing'], datatypes.STRING_SUBTYPES['nothing'], datatypes.HEADER_TYPES['string'])
                ################################
                        #END(3)
                ################################
            
            ###################################################################
            ###################################################################

            if header_type[i] == datatypes.HEADER_TYPES['date']:
                if re.search(r'( HH:mm:ss)$', format_type[i]):
                    if not chop_zero[i] == 2:
                        if re.search(r'( 00:00:00)$', tempDate):
                            chop_zero[i] = 1
                        else:
                            chop_zero[i] = 2
                            
#                if (temp != 'nan' and temp != 'nat' and temp != 'na') or temp == "":
                if (temp != 'nan' and temp != 'nat') or temp == "":
                    single_row.append(tempDate)
                else:
                    single_row.append("")
                    
            elif header_type[i] == datatypes.HEADER_TYPES['number']:
#                if (temp != 'nan' and temp != 'nat' and temp != 'na') or temp == "":
                if (temp != 'nan' and temp != 'nat') or temp == "":
                    single_row.append(temp)
                else:
                    single_row.append("")
                    
            else:
#                if (temp != 'nan' and temp != 'nat' and temp != 'na') or temp == "":
                if (temp != 'nan' and temp != 'nat') or temp == "":
                    single_row.append(element)
                else:
                    single_row.append("")
        
        i += 1
    
    return [temp_bool, single_row, header_type, sub_type, format_type, chop_zero]

def isExactlyOneVisibleSheet(filePath):

    tempBool = True
    
    xls = pd.ExcelFile(filePath)
    sheets = xls.book.sheets()
    
    visible = 0
    sheet_name = ""
    for sheet in sheets:
        if sheet.visibility == 0:
            visible += 1
            sheet_name += sheet.name
            if visible > 1:
                tempBool = False
                break
                
    return [tempBool, sheet_name]

def useCheckMeta_Empty(fileName, status):
    
    var = f_load.checkMeta_Empty(fileName)
    
    if var:
        f_load.db.datasets.update_one({'files.orgFileId' : fileName}, {'$set': {'status': status}})
        f_load.post_result(fileName)
        f_load.update_mergeStatus(fileName, status)
        # post_mergeStatus(fileName)
    else:
        f_load.update_mergeStatus(fileName, status)
        f_load.post_mergeStatus(fileName)

def createEmptyIndexDictList(row, first_data_index):
    len_row = len(row)
    emptyColumnHeaderArray = []
    for i in range(len_row - first_data_index):
        temp = row[first_data_index::][i].replace(" ", "").lower()
        if isCellEmpty(temp):
            emptyColumnHeaderArray.append({'colId': i, 'orgColId': i, 'fixed': False})
        else:
            pass
        
    return emptyColumnHeaderArray

def createDupIndexDictList(row, first_data_index):
    distinct_list = []
    indices = []
    duplicateHeaderArray = []
    len_row = len(row)
    
    for i in range(len_row - first_data_index):
        temp = row[first_data_index::][i].replace(" ", "").lower()
        if not isCellEmpty(temp):
            if not temp in distinct_list:
                distinct_list.append(temp)
                indices.append([i])
            else:
                indices[distinct_list.index(temp)].append(i)
        else:
            pass
    
    k = -1
    for i in range(len(distinct_list)):
        if (len(indices[i])) > 1:
            k += 1
            duplicateHeaderArray.append({'duplicateList':[], 'value': distinct_list[i], 'fixedLeastOne': False})
            for j in range(len(indices[i])):
                duplicateHeaderArray[k]['duplicateList'].append({'duplicateColId': indices[i][j], 'orgColId': indices[i][j], 'colValue': distinct_list[i], 'fixed': False})
        else:
            pass
        
    return duplicateHeaderArray

def emptyHeadersExists(row, first_data_index):
    exists = False
    
    for i in range(len(row) - first_data_index):
        temp = row[first_data_index::][i].replace(" ", "").lower()
        if isCellEmpty(temp):
            exists = True
            break
    
    return exists
    
def duplicateHeadersExists(row):
    exists = False
    distinct_list = []
    
    for header in row:
       temp = header.replace(" ","").lower()
       if not isCellEmpty(temp):
           if not temp in distinct_list:
               distinct_list.append(temp)
           else:
               exists = True
               break
       else:
           pass
       
    return exists

def get_startingPosition(raw_list, len_raw_list, first_nonemptyIndex, headersRow_index):
    
    first_data_index = first_nonemptyIndex
    
    if first_data_index == 0:
        pass
    else:
        for row_num in range (headersRow_index + 1, len_raw_list):
            row = list(map(str, raw_list[row_num]))
            for i in range(first_data_index):
                if not isCellEmpty(row[i].replace(" ", "").lower()):
                    first_data_index = i
                    if first_data_index == 0:
                        break
                    
    return first_data_index

def errorCorrection(raw_list, fileName, first_data_index, headersRow_index):
    
    len_raw_list = len(raw_list)
    j = 0
    data = []
    error_code = 0
    emptyColumnHeaderArray = []
    duplicateHeaderArray = []
    error_message = None
    
    for i in range(headersRow_index,len_raw_list):
        row = list(map(str, raw_list[i]))
        
        if not isEveryElementEmpty(row)[0]:
            data.append(row[first_data_index::])
            j += 1
            if j == 1:
                len_header_list = len(data[0])
                for i in range(len_header_list):
                    if isCellEmpty(data[0][i].lower().replace(" ","")):
                        data[0][i] = ""
                if emptyHeadersExists(row, first_data_index):
                    emptyColumnHeaderArray += createEmptyIndexDictList(row, first_data_index)
                    error_code += 1
                if duplicateHeadersExists(row):
                    duplicateHeaderArray += createDupIndexDictList(row, first_data_index)
                    error_code += 2
            else:
                pass
        else:
            pass
        
        if j == 31:
            break
        
    if error_code == 1:
        error_message = "Error! Empty Header(s)"
    elif error_code == 2:
        error_message = "Error! Duplicate Header(s)"
    elif error_code == 3:
        error_message = "Error! Duplicate and Empty Headers"
        
    f_load.update_fileErrorStatus(fileName, 'errorInfo', error_message)
    
    return [data, emptyColumnHeaderArray, duplicateHeaderArray]

def post_correction(fileName, correction_list):

    dataset1 = f_load.db.datasets.find_one({'files.orgFileId': fileName })
    dataset_Id = dataset1['_id']

    body = {
        "userId": dataset1['userId'],
        "componentId": "errorCorrectionCom",
        "mergeStatus": f_load.get_mergeStatus(fileName),
        "dataSetId": str(f_load.bson.objectid.ObjectId(dataset_Id)),
        "fileId": f_load.getFileId(fileName),
        "fileName": f_load.get_fileName(fileName),
        "data": json.dumps(correction_list[0]),
        "emptyColumnHeaderArray": json.dumps(correction_list[1]),
        "duplicateHeaderArray": json.dumps(correction_list[2])
    }
    f_load.requests.post(config.SOCKET['upload_correction'], body)

def identify_firstIndex(raw_list):
    
    len_raw_list = len(raw_list)
    identified = False
    error_code = 0
    headersRow_index = 0
    first_nonemptyIndex = 0
    
    for i in range(len_raw_list):
        row = list(map(str, raw_list[i]))
        if not isEveryElementEmpty(row)[0]:
            for k in range(len(row)):
                temp = row[k].replace(" ", "").lower()
                if not isCellEmpty(temp):
                    first_nonemptyIndex = k
                    identified = True
                    break
            if identified == True:
                headersRow_index = i
                break
        else:pass
    
    if headersRow_index == len_raw_list - 1:
        error_code += 1
        
    if len_raw_list == 0:
        error_code += 2
    
    return first_nonemptyIndex, headersRow_index, error_code

def correct_emptyAndDuplicateHeaders(raw_list, first_data_index, headersRow_index, emptyHeadersCorrected, duplicateHeadersCorrected):
    
    raw_list[headersRow_index] = list(map(str, raw_list[headersRow_index]))
    
    if len(emptyHeadersCorrected) > 0:
        for i in range(len(emptyHeadersCorrected)):
            if 'colValue' in emptyHeadersCorrected[i]:
                raw_list[headersRow_index][emptyHeadersCorrected[i]['orgColId'] + first_data_index] = emptyHeadersCorrected[i]['colValue']
            else:
                pass
    
    if len(duplicateHeadersCorrected) > 0:
        for i in range(len(duplicateHeadersCorrected)):
            for j in range(len(duplicateHeadersCorrected[i]['duplicateList'])):
                raw_list[headersRow_index][duplicateHeadersCorrected[i]['duplicateList'][j]['orgColId'] + first_data_index] = duplicateHeadersCorrected[i]['duplicateList'][j]['colValue']

    return raw_list


def rev_sortDeletedHeaders(deletedHeaders, first_data_index):
    
    deletedHeaders_asListRev = []
    
    if len(deletedHeaders) > 0:
        deletedHeaders = sorted(deletedHeaders, key = lambda i: i['orgColId'], reverse = True)
        for i in deletedHeaders:
            deletedHeaders_asListRev.append(i['orgColId'] + first_data_index)
        
    return deletedHeaders_asListRev

def neglect_indices(lst, deletedHeaders_asListRev):
    
    if len(deletedHeaders_asListRev) > 0:
        for i in deletedHeaders_asListRev:
            del lst[i]
        
    return lst

#set the userCorrectionStatus of the file if there is any wrong to be correct
def set_userCorrectionStatus(fileName,status):

    client = f_load.MongoClient(config.MONGO['client'])
    db = client[f_load.database_Name]

    dataset1 = f_load.getDataSet(fileName)
    files = dataset1['files']

    for i in files:
        if  (i['orgFileId'] == fileName):
            i['userCorrectionStatus'] = status
         
    db.datasets.update_one({'files.orgFileId' : fileName}, {'$set': {'files': files}})

#check the userCorrectionStatus of the file.
def get_userCorrectionStatus(fileName):
    
    
    print("file name", fileName)
    dataset1 = f_load.getDataSet(fileName)
    print("dataset1", dataset1)
    if not dataset1 == None:
        files = dataset1['files']
    
        for i in files:
            if (i['orgFileId'] == fileName):
                if 'userCorrectionStatus' in i:
                    return (i['userCorrectionStatus'])
                else:
                    return None
    else: return "" 
    
def get_errorsTuple(fileName):

    dataset1 = f_load.getDataSet(fileName)
    files = dataset1['files']
    
    print ("within 'get_errorsTuple'...")
    
    for i in files:
        if  (i['orgFileId'] == fileName):
            file_object = i
            print ("file_object", file_object)

    deleted_headers = file_object['errors']['deletedHeaders']
    empty_headers = file_object['errors']['emptyHeadersCorrected']
    duplicate_headers = file_object['errors']['duplicateHeadersCorrected']
    
    print ("deleted_headers")
    print (deleted_headers)
    print ("empty_headers")
    print (empty_headers)
    print ("duplicate_headers")
    print (duplicate_headers)

    return deleted_headers, empty_headers, duplicate_headers

def csv_transform(path, user_correction_status, fileName):
    
    if os.path.getsize(path) < 50000000:
        fileweight = 1
    else:
        fileweight = 2
    
    tem = True
    time1 = time.time()
    
    merge_status = None
    while(tem):
        time2 = time.time()
        try:
            merge_status = f_load.get_mergeStatus(fileName)
            tem = False
        except:
            if time2 - time1 > 30*fileweight:
                tem = False
                useCheckMeta_Empty(fileName, config.DATABASE_STATES['process_fail'])
                sendemail(fileName, "Process Failed after waiting for front end to update DB")
                return [path]


    run = 0
    i = 0
    exception = 0
    data_presence = False
    
    if merge_status == config.DATABASE_STATES['upload_success'] or merge_status == config.DATABASE_STATES['process_success']:
        
        useCheckMeta_Empty(fileName, config.DATABASE_STATES['process_start'])
        run = 1
    
    if run == 1:
        try:
            raw_list = pd.read_csv(path, header = None).values.tolist()
            
            tuple_identify_firstIndex = identify_firstIndex(raw_list)
            first_data_index = get_startingPosition(raw_list, len(raw_list), tuple_identify_firstIndex[0], tuple_identify_firstIndex[1])
            
            if user_correction_status == config.USER_CORRECTION_STATUS['none']:
                if tuple_identify_firstIndex[2] == 0:
                    
                    deletedHeaders_asListRev = []
                    data_presence = True
                    
            else:##if user_correction_status == config.USER_CORRECTION_STATUS['success']
                errors_tuple = get_errorsTuple(fileName)
        
                deletedHeaders = errors_tuple[0]
                emptyHeadersCorrected = errors_tuple[1]
                duplicateHeadersCorrected = errors_tuple[2]
                deletedHeaders_asListRev = rev_sortDeletedHeaders(deletedHeaders, tuple_identify_firstIndex[0])
                
                raw_list = correct_emptyAndDuplicateHeaders(raw_list, first_data_index, tuple_identify_firstIndex[1], emptyHeadersCorrected, duplicateHeadersCorrected)
            
            headers_list = []
            raw_data_list = []
                    
            if data_presence == True or user_correction_status == config.USER_CORRECTION_STATUS['success']:
                
                for row in raw_list:
                    row = list(map(str, row))
                    if i == 0:
                        list_isEveryElementEmpty = isEveryElementEmpty(row)
                        
                        if not list_isEveryElementEmpty[0]:
                            list_isExistsDuplicate = isExistsDuplicate(row,  deletedHeaders_asListRev, first_data_index, user_correction_status)
                            
                            if user_correction_status == config.USER_CORRECTION_STATUS['none']:
                                
                                if list_isExistsDuplicate[0] or list_isExistsDuplicate[2]:
                                    
                                    set_userCorrectionStatus(fileName,config.USER_CORRECTION_STATUS['pending'])
                                    shutil.copy2(path, config.WATCH_PATHS['temp'] + path.split("/")[-1])
                                    print("FIRST copy: "+path.split("/")[-1])
                                    post_correction(fileName, errorCorrection(raw_list, fileName, first_data_index, tuple_identify_firstIndex[1]))
                                    break
                                else:
                                    
                                    if len(list_isExistsDuplicate[1]) > 100:
                                        
                                        f_load.update_fileErrorStatus(fileName, 'errorInfo',  errors.PROCESS_ERRORS['column_limit_exceeded'])
                                        break
                                    else:
                                        
                                        list_generateLists = generateLists(row)
                                        headers_list += list_isExistsDuplicate[1]
                                        i += 1
                                        
                            else:#if user_correction_status == config.USER_CORRECTION_STATUS['success']
                                
                                if list_isExistsDuplicate[0] and list_isExistsDuplicate[2]:
                                    f_load.update_fileErrorStatus(fileName, 'errorInfo', errors.PROCESS_ERRORS['empty_and_duplicate_header'])
                                    break
                                elif list_isExistsDuplicate[0]:
    
                                    f_load.update_fileErrorStatus(fileName, 'errorInfo',  errors.PROCESS_ERRORS['duplicate_header'])
                                    break
                                elif list_isExistsDuplicate[2]:
    
                                    f_load.update_fileErrorStatus(fileName, 'errorInfo',  errors.PROCESS_ERRORS['empty_header'])
                                    break
                                else:
                                    if len(list_isExistsDuplicate[1]) > 100:
                                        f_load.update_fileErrorStatus(fileName, 'errorInfo',  errors.PROCESS_ERRORS['column_limit_exceeded'])
                                        break
                                    else:
                                        list_generateLists = generateLists(row)
                                        headers_list += list_isExistsDuplicate[1]
                                        i += 1   
                                
                        else:
                            continue
                    
                    elif i == 1:
                        
                        list_isEveryElementEmpty = isEveryElementEmpty(row)
                        if list_isEveryElementEmpty[0]:
                            continue
                        
                        else:
                            
                            list_isAgainstHeaderIndices = isAgainstHeaderIndices(row, list_generateLists[0], list_generateLists[1], list_generateLists[2], headers_list, deletedHeaders_asListRev, first_data_index, list_generateLists[3], user_correction_status)
                        
                        if list_isAgainstHeaderIndices[0]:
                            f_load.update_fileErrorStatus(fileName, 'errorInfo',  errors.PROCESS_ERRORS['mismatching_data&header_index'])
                            i += 1
                            break
                        
                        raw_data_list.append(list_isAgainstHeaderIndices[1]\
                                         [first_data_index::])
                    
            else:
                if tuple_identify_firstIndex[2] == 1:
                    f_load.update_fileErrorStatus(fileName, 'errorInfo',  errors.PROCESS_ERRORS['only headers'])
                else:
                    f_load.update_fileErrorStatus(fileName, 'errorInfo',  errors.PROCESS_ERRORS['empty_file'])
                    
        except Exception as instance:
            if str(type(instance)) == "<class 'UnicodeDecodeError'>":
                f_load.update_fileErrorStatus(fileName, 'errorInfo',  "Invalid characters like ¡,¢,£,¤,¥,§,©,¼,Æ,Ø,ß,� are found. Please clean and upload")
            else:
                f_load.update_fileErrorStatus(fileName, 'errorInfo',  str(type(instance)) + str(instance))

            useCheckMeta_Empty(fileName, config.DATABASE_STATES['process_fail'])
            exception += 1
            
        if exception == 0:
            if i == 1:
                list_cleanData = cleanData(raw_data_list, list_isAgainstHeaderIndices[2][first_data_index::], list_isAgainstHeaderIndices[3][first_data_index::], list_isAgainstHeaderIndices[4][first_data_index::], list_isAgainstHeaderIndices[5][first_data_index::])
                
                headerid = headers_list
                dataid = list_cleanData[0]
                typeid = list_cleanData[1]
                subid = list_cleanData[2]
                formatid = list_cleanData[3]
                    
                useCheckMeta_Empty(fileName, config.DATABASE_STATES['process_success'])
                return [path,headerid,typeid,dataid,subid,formatid]
                
            else:
                useCheckMeta_Empty(fileName, config.DATABASE_STATES['process_fail'])
                sendemail(fileName, "Process Failed")
                return [path]
        else:
            return [path]
    else:
        sendemail(fileName, "Process Failed due to status in DB is wrong")
        return [path]

def xls_transform(path, user_correction_status, fileName):
    
    if os.path.getsize(path) < 50000000:
        fileweight = 1
    else:
        fileweight = 2
    
    tem = True
    time1 = time.time()
    
    merge_status = None
    while(tem):
        time2 = time.time()
        try:
            merge_status = f_load.get_mergeStatus(fileName)
            tem = False
        except:
            if time2 - time1 > 30*fileweight:
                tem = False
                useCheckMeta_Empty(fileName, config.DATABASE_STATES['process_fail'])
                sendemail(fileName, "Process Failed after waiting for front end to update DB")
                return [path]


    run = 0
    i = 0
    exception = 0
    data_presence = False
    
    if merge_status == config.DATABASE_STATES['upload_success'] or merge_status == config.DATABASE_STATES['process_success']:
        
        useCheckMeta_Empty(fileName, config.DATABASE_STATES['process_start'])
        run = 1
        
    if run == 1:
        try:
            list_isExactlyOneVisibleSheet = isExactlyOneVisibleSheet(path)
            
            if not list_isExactlyOneVisibleSheet[0]:
                f_load.update_fileErrorStatus(fileName, 'errorInfo',  errors.PROCESS_ERRORS['multiple_visible_sheets'])
            else:
                wb = xlrd.open_workbook(path, formatting_info = True)
        
                xl_sheet = wb.sheet_by_name(list_isExactlyOneVisibleSheet[1])
                
                unhidden_cols = []
                for xl_col in range(xl_sheet.ncols):
                    if xl_sheet.colinfo_map[xl_col].hidden == False:
                        unhidden_cols.append(xl_col)
                        
                raw_list = pd.read_excel(path, sheet_name = list_isExactlyOneVisibleSheet[1], header = None, usecols = unhidden_cols).values.tolist()
                
                tuple_identify_firstIndex = identify_firstIndex(raw_list)
                first_data_index = get_startingPosition(raw_list, len(raw_list), tuple_identify_firstIndex[0], tuple_identify_firstIndex[1])
                
                if user_correction_status == config.USER_CORRECTION_STATUS['none']:
                    if tuple_identify_firstIndex[2] == 0:
                        
                        deletedHeaders_asListRev = []
                        data_presence = True
                    
                else:##if user_correction_status == config.USER_CORRECTION_STATUS['success']
                    errors_tuple = get_errorsTuple(fileName)
            
                    deletedHeaders = errors_tuple[0]
                    emptyHeadersCorrected = errors_tuple[1]
                    duplicateHeadersCorrected = errors_tuple[2]
                    deletedHeaders_asListRev = rev_sortDeletedHeaders(deletedHeaders, first_data_index)
                    
                    raw_list = correct_emptyAndDuplicateHeaders(raw_list, first_data_index, tuple_identify_firstIndex[1], emptyHeadersCorrected, duplicateHeadersCorrected)
                
                headers_list = []
                raw_data_list = []
                if data_presence == True or user_correction_status == config.USER_CORRECTION_STATUS['success']:
                    list_row_nums = [x for x in range (xl_sheet.nrows)]
                    
                    for rowpd, rowxl in zip(raw_list, list_row_nums):
                        row = list(map(str, rowpd))
                        
                        if xl_sheet.rowinfo_map[rowxl].hidden == False:
                            if i == 0:
                                list_isEveryElementEmpty = isEveryElementEmpty(row)
                                if not list_isEveryElementEmpty[0]:
                                    list_isExistsDuplicate = isExistsDuplicate(row,  deletedHeaders_asListRev, first_data_index, user_correction_status)
                                    
                                    if user_correction_status == config.USER_CORRECTION_STATUS['none']:
                                        if list_isExistsDuplicate[0] or list_isExistsDuplicate[2]:
                                            set_userCorrectionStatus(fileName,config.USER_CORRECTION_STATUS['pending'])
                                            shutil.copy2(path, config.WATCH_PATHS['temp'] + path.split("/")[-1])
                                            print("FIRST copy: "+path.split("/")[-1])
                                            post_correction(fileName, errorCorrection(raw_list, fileName, first_data_index, tuple_identify_firstIndex[1]))
                                            break
                                        else:
                                            if len(list_isExistsDuplicate[1]) > 100:
                                                f_load.update_fileErrorStatus(fileName, 'errorInfo',  errors.PROCESS_ERRORS['column_limit_exceeded'])
                                                break
                                            else:
                                                list_generateLists = generateLists(row)
                                                headers_list += list_isExistsDuplicate[1]
                                                i += 1
                                                
                                    else:#if user_correction_status == config.USER_CORRECTION_STATUS['success']
                                        if list_isExistsDuplicate[0] and list_isExistsDuplicate[2]:
                                            f_load.update_fileErrorStatus(fileName, 'errorInfo', errors.PROCESS_ERRORS['empty_and_duplicate_header'])
                                            break
                                        elif list_isExistsDuplicate[0]:
            
                                            f_load.update_fileErrorStatus(fileName, 'errorInfo',  errors.PROCESS_ERRORS['duplicate_header'])
                                            break
                                        elif list_isExistsDuplicate[2]:
            
                                            f_load.update_fileErrorStatus(fileName, 'errorInfo',  errors.PROCESS_ERRORS['empty_header'])
                                            break
                                        else:
                                            if len(list_isExistsDuplicate[1]) > 100:
                                                f_load.update_fileErrorStatus(fileName, 'errorInfo',  errors.PROCESS_ERRORS['column_limit_exceeded'])
                                                break
                                            else:
                                                list_generateLists = generateLists(row)
                                                headers_list += list_isExistsDuplicate[1]
                                                i += 1
                                else:
                                    continue
                            
                            elif i == 1:
                                list_isEveryElementEmpty = isEveryElementEmpty(row)
                                if list_isEveryElementEmpty[0]:
                                    continue
                                
                                else:
                                    list_isAgainstHeaderIndices = isAgainstHeaderIndices(row, list_generateLists[0], list_generateLists[1], list_generateLists[2], headers_list, deletedHeaders_asListRev, first_data_index, list_generateLists[3], user_correction_status)
                                
                                if list_isAgainstHeaderIndices[0]:
                                    f_load.update_fileErrorStatus(fileName, 'errorInfo',  errors.PROCESS_ERRORS['mismatching_data&header_index'])   
                                    i += 1
                                    break
                                
                                raw_data_list.append(list_isAgainstHeaderIndices[1]\
                                                 [first_data_index::])
                        
                        else:
                            pass
                        
                else:
                    if tuple_identify_firstIndex[2] == 1:
                        f_load.update_fileErrorStatus(fileName, 'errorInfo',  errors.PROCESS_ERRORS['only headers'])
                    else:
                        f_load.update_fileErrorStatus(fileName, 'errorInfo',  errors.PROCESS_ERRORS['empty_file'])
                    
        except Exception as instance:

            f_load.update_fileErrorStatus(fileName, 'errorInfo',  str(type(instance)) + str(instance))
            useCheckMeta_Empty(fileName, config.DATABASE_STATES['process_fail'])
            exception += 1

        if exception == 0:
            if i == 1:
                list_cleanData = cleanData(raw_data_list, list_isAgainstHeaderIndices[2][first_data_index::], list_isAgainstHeaderIndices[3][first_data_index::], list_isAgainstHeaderIndices[4][first_data_index::], list_isAgainstHeaderIndices[5][first_data_index::])
                
                headerid = headers_list
                dataid = list_cleanData[0]
                typeid = list_cleanData[1]
                subid = list_cleanData[2]
                formatid = list_cleanData[3]
                    
                useCheckMeta_Empty(fileName, config.DATABASE_STATES['process_success'])
    
                return [path,headerid,typeid,dataid,subid,formatid]
                
            else:
                useCheckMeta_Empty(fileName, config.DATABASE_STATES['process_fail'])
                sendemail(fileName, "Process Failed")
                return [path]
        else:
            return [path]
    else:
        sendemail(fileName, "Process Failed due to status in DB is wrong")
        return [path]   
    
def xlsx_transform(path, user_correction_status, fileName):

    if os.path.getsize(path) < 50000000:
        fileweight = 1
    else:
        fileweight = 2
    
    tem = True
    time1 = time.time()
    
    merge_status = None
    while(tem):
        time2 = time.time()
        try:
            merge_status = f_load.get_mergeStatus(fileName)
            tem = False
        except:
            if time2 - time1 > 30*fileweight:
                tem = False
                useCheckMeta_Empty(fileName, config.DATABASE_STATES['process_fail'])
                sendemail(fileName, "Process Failed after waiting for front end to update DB")
                return [path]


    run = 0
    i = 0
    exception = 0
    data_presence = False
    
    if merge_status == config.DATABASE_STATES['upload_success'] or merge_status == config.DATABASE_STATES['process_success']:
        
        useCheckMeta_Empty(fileName, config.DATABASE_STATES['process_start'])
        run = 1
        
        if run == 1:
            try:
                list_isExactlyOneVisibleSheet = isExactlyOneVisibleSheet(path)
                
                if not list_isExactlyOneVisibleSheet[0]:
                    f_load.update_fileErrorStatus(fileName, 'errorInfo',  errors.PROCESS_ERRORS['multiple_visible_sheets'])
                else:
                    wb = openpyxl.load_workbook(path)
                    ws = wb[list_isExactlyOneVisibleSheet[1]]
                    
                    unhidden_cols = ""  
                    for column in ws.iter_cols(max_col = ws.max_column): 
                        if ws.column_dimensions[column[0].column].hidden == False:
                            unhidden_cols += column[0].column + ","
                    
                    raw_list = pd.read_excel(path, sheet_name = list_isExactlyOneVisibleSheet[1], header = None, usecols = unhidden_cols).values.tolist()
                    
                    tuple_identify_firstIndex = identify_firstIndex(raw_list)
                    first_data_index = get_startingPosition(raw_list, len(raw_list), tuple_identify_firstIndex[0], tuple_identify_firstIndex[1])
                
                    if user_correction_status == config.USER_CORRECTION_STATUS['none']:
                        if tuple_identify_firstIndex[2] == 0:
                            
                            deletedHeaders_asListRev = []
                            data_presence = True
                        
                    else:##if user_correction_status == config.USER_CORRECTION_STATUS['success']
                        errors_tuple = get_errorsTuple(fileName)
                
                        deletedHeaders = errors_tuple[0]
                        emptyHeadersCorrected = errors_tuple[1]
                        duplicateHeadersCorrected = errors_tuple[2]
                        deletedHeaders_asListRev = rev_sortDeletedHeaders(deletedHeaders, first_data_index)
                        
                        raw_list = correct_emptyAndDuplicateHeaders(raw_list, first_data_index, tuple_identify_firstIndex[1], emptyHeadersCorrected, duplicateHeadersCorrected)
                    
                    
                    headers_list = []
                    raw_data_list = []
                    if data_presence == True or user_correction_status == config.USER_CORRECTION_STATUS['success']:
                    
                        for rowpd, rowws in zip(raw_list, ws):
                            
                            row = list(map(str, rowpd))
                            if ws.row_dimensions[rowws[0].row].hidden == False:
                                
                                if i == 0:
                                    list_isEveryElementEmpty = isEveryElementEmpty(row)
                                    
                                    if not list_isEveryElementEmpty[0]:
                                        list_isExistsDuplicate = isExistsDuplicate(row,  deletedHeaders_asListRev, first_data_index, user_correction_status)
                                        
                                        if user_correction_status == config.USER_CORRECTION_STATUS['none']:
                                            if list_isExistsDuplicate[0] or list_isExistsDuplicate[2]:
                                                set_userCorrectionStatus(fileName,config.USER_CORRECTION_STATUS['pending'])
                                                shutil.copy2(path, config.WATCH_PATHS['temp'] + path.split("/")[-1])
                                                print("FIRST copy: "+path.split("/")[-1])
                                                post_correction(fileName, errorCorrection(raw_list, fileName, first_data_index, tuple_identify_firstIndex[1]))
                                                break
                                            else:
                                                if len(list_isExistsDuplicate[1]) > 100:
                                                    f_load.update_fileErrorStatus(fileName, 'errorInfo',  errors.PROCESS_ERRORS['column_limit_exceeded'])
                                                    break
                                                else:
                                                    list_generateLists = generateLists(row)
                                                    headers_list += list_isExistsDuplicate[1]
                                                    i += 1
                                                    
                                        else:#if user_correction_status == config.USER_CORRECTION_STATUS['success']
                                            if list_isExistsDuplicate[0] and list_isExistsDuplicate[2]:
                                                f_load.update_fileErrorStatus(fileName, 'errorInfo', errors.PROCESS_ERRORS['empty_and_duplicate_header'])
                                                break
                                            elif list_isExistsDuplicate[0]:
                
                                                f_load.update_fileErrorStatus(fileName, 'errorInfo',  errors.PROCESS_ERRORS['duplicate_header'])
                                                break
                                            elif list_isExistsDuplicate[2]:
                
                                                f_load.update_fileErrorStatus(fileName, 'errorInfo',  errors.PROCESS_ERRORS['empty_header'])
                                                break
                                            else:
                                                if len(list_isExistsDuplicate[1]) > 100:
                                                    f_load.update_fileErrorStatus(fileName, 'errorInfo',  errors.PROCESS_ERRORS['column_limit_exceeded'])
                                                    break
                                                else:
                                                    list_generateLists = generateLists(row)
                                                    headers_list += list_isExistsDuplicate[1]
                                                    i += 1
                                    else:
                                        continue
                                
                                elif i == 1:
                                    list_isEveryElementEmpty = isEveryElementEmpty(row)
                                    
                                    if list_isEveryElementEmpty[0]:
                                        continue
                                    
                                    else:
                                        list_isAgainstHeaderIndices = isAgainstHeaderIndices(row, list_generateLists[0], list_generateLists[1], list_generateLists[2], headers_list, deletedHeaders_asListRev, first_data_index, list_generateLists[3], user_correction_status)
                                    
                                    if list_isAgainstHeaderIndices[0]:
                                        f_load.update_fileErrorStatus(fileName, 'errorInfo',  errors.PROCESS_ERRORS['mismatching_data&header_index'])
                                        i += 1
                                        break
                                    
                                    raw_data_list.append(list_isAgainstHeaderIndices[1]\
                                                        [first_data_index::])
                                
                            else:
                                pass
                    
                    else:
                        if tuple_identify_firstIndex[2] == 1:
                            f_load.update_fileErrorStatus(fileName, 'errorInfo',  errors.PROCESS_ERRORS['only headers'])
                        else:
                            f_load.update_fileErrorStatus(fileName, 'errorInfo',  errors.PROCESS_ERRORS['empty_file'])       
                       
            except Exception as instance:
    
                f_load.update_fileErrorStatus(fileName, 'errorInfo',  str(type(instance)) + str(instance))
                useCheckMeta_Empty(fileName, config.DATABASE_STATES['process_fail'])
                exception += 1
            
            if exception == 0:
                if i == 1:
                    list_cleanData = cleanData(raw_data_list, list_isAgainstHeaderIndices[2][first_data_index::], list_isAgainstHeaderIndices[3][first_data_index::], list_isAgainstHeaderIndices[4][first_data_index::], list_isAgainstHeaderIndices[5][first_data_index::])
                    
                    headerid = headers_list
                    dataid = list_cleanData[0]
                    typeid = list_cleanData[1]
                    subid = list_cleanData[2]
                    formatid = list_cleanData[3]
                        
                    useCheckMeta_Empty(fileName, config.DATABASE_STATES['process_success'])
        
                    return [path,headerid,typeid,dataid,subid,formatid]
                    
                else:
                    useCheckMeta_Empty(fileName, config.DATABASE_STATES['process_fail'])
                    sendemail(fileName, "Process Failed")
                    return [path]
            else:
                return [path]
        else:
            sendemail(fileName, "Process Failed due to status in DB is wrong")
            return [path] 
        
def sendemail(subject, content):
    msg = EmailMessage()
    msg.set_content(content)
    msg['Subject'] = subject
    msg['From'] = mail.MAIL['from']
    msg['To'] = mail.MAIL['to']

    server = smtplib.SMTP(mail.MAIL['smtp'])
    server.starttls()
    server.login(mail.MAIL['username'],mail.MAIL['password'])
    server.send_message(msg)
    server.quit()